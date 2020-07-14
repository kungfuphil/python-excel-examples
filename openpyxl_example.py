# This module can read and write to files with the .xlsx extension
from openpyxl import Workbook, load_workbook

destination_spreadsheet_path = "destination_spreadsheet.xlsx"
destination_workbook = load_workbook(destination_spreadsheet_path)

# This is the first worksheet available
destination_worksheet = destination_workbook.worksheets[0]

destination_worksheet.cell(row=1, column=1).value = "This is cell A1"
destination_workbook.save(destination_spreadsheet_path)

# You don't have to save after every cell. This is just to show
# that you can continue to append to the same workbook.
destination_worksheet.cell(row=1, column=2).value = "This is cell B1"
destination_workbook.save(destination_spreadsheet_path)