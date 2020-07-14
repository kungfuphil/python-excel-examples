# This module is to read files with the .xls extension
from xlrd import open_workbook

# This module can read and write to files with the .xlsx extension
from openpyxl import Workbook, load_workbook

source_spreadsheet_path = "source_spreadsheet.xls"
source_workbook = open_workbook(source_spreadsheet_path)

# This is the first worksheet available
source_worksheet = source_workbook.sheet_by_index(0)

# Columns in xlrd start at index 0
source_max_columns = source_worksheet.ncols

# Rows in xlrd start at index 1
source_max_rows = source_worksheet.nrows

destination_spreadsheet_path = "destination_spreadsheet.xlsx"
destination_workbook = load_workbook(destination_spreadsheet_path)

destination_worksheet = destination_workbook.worksheets[0]

# Rows in openpyxl start at index 1, columns also start at index 1
destination_max_rows = destination_worksheet.max_row

# If the first cell of the first row is empty, start writing on the first row,
# otherwise start in the max row count + 1, which should be the next empty row.
destination_starting_row = 1 if destination_worksheet.cell(1, 1).value is None else destination_max_rows + 1

# Iterate through the rows of the source file and copy them
# to the destination file.
# for i in range(1, source_max_rows): # If you want to skip the header, use this line instead of the next
for i in range(source_max_rows):
    row = source_worksheet.row_values(i)
    for j, source_cell_value in enumerate(row):
        row_num = destination_starting_row + i
        destination_worksheet.cell(row=row_num, column=j+1).value = source_cell_value

destination_workbook.save(destination_spreadsheet_path)